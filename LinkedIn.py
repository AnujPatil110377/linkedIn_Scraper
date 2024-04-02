from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
import time
import undetected_chromedriver as uc
import os

def scrape_linkedin(company_name, job_keywords, driver):
    # Construct the search query URL
    search_query = '+'.join(company_name.split()) + '+' + job_keywords + '+linkedin'
    search_url = f"https://www.google.com/search?q={search_query}"

    # Perform Google search
    driver.get(search_url)
    time.sleep(2)  # Add a delay to ensure page load

    # Parse search results
    soup = BeautifulSoup(driver.page_source, 'html.parser')
    search_results = soup.find_all('div', class_='g')

    scraped_data = []
    scraped_urls = set()  # Keep track of scraped LinkedIn URLs

    for result in search_results:
        try:
            name_elem = result.find('h3', class_='LC20lb')
            name_parts = name_elem.text.strip().split(' - ')
            name = name_parts[0] if name_parts else None

            position_elem = result.find('div', class_='LEwnzc')
            position_parts = position_elem.text.strip().split(' · ')
            position = position_parts[1] if position_parts else None

            linkedin_url_elem = result.find('a', href=True)
            linkedin_url = linkedin_url_elem['href'] if linkedin_url_elem else None

            # Check if the LinkedIn URL has already been scraped
            if linkedin_url and linkedin_url not in scraped_urls:
                scraped_urls.add(linkedin_url)  # Add URL to set
                if name and position:
                    scraped_data.append((name, position, linkedin_url))
        except Exception as e:
            print(f"Error parsing profile: {e}")

    return scraped_data

def main():
    # Load company names from Excel sheet
    wb = load_workbook('company.xlsx')
    sheet = wb.active
    companies = [cell.value for row in sheet.iter_rows() for cell in row if cell.value]

    # Specify job keywords as a list of individual keywords
    job_keywords = ["Campus recruitment", "talent acquisition", "-hr"]

    # Initialize Chrome WebDriver only once using undetected_chromedriver
    options = uc.ChromeOptions()
    driver = uc.Chrome(options=options)

    # Create a new Workbook for the LinkedIn profiles
    wb_new = Workbook()
    sheet_new = wb_new.active
    sheet_new.append([" ", " ", " ", " ", "position", " ", "name", " ", "linkedin_url"])

    # Scrape LinkedIn profiles for each company and keyword
    for company in companies:
        sheet_new.append([company])
        sheet_new.append([" "])
        for keyword in job_keywords:
            print(f"Scraping LinkedIn for {company} - {keyword}...")
            data = scrape_linkedin(company, keyword, driver)
            for name, position, linkedin_url in data:
                sheet_new.append([" ", " ", " ", " ", name, " ", keyword, " ", linkedin_url])
            print(f"Found {len(data)} profiles for {company} - {keyword}")

    # Close the WebDriver after all searches are done
    driver.quit()

    # Save the LinkedIn profiles to a new Excel file
    wb_new.save('linkedin_profiles.xlsx')

if __name__ == "__main__":
    main()
